from openpyxl import Workbook, load_workbook
import os, datetime
from my_logging import Log

# class Writer:
#     def __init__(self):
#         self.wb = None
#         self.ws = None
#
#     def create(self, name_filename, name_sheet):
#         # Удаляю стандартный лист
#
#         # Создаю новый
#         filename = '{}.xlsx'.format(name_filename)
#         if os.path.isfile(filename):
#             print('File exist \n')
#             self.wb = load_workbook(filename=filename)
#             self.ws = self.wb.active
#         else:
#             print('New file \n')
#             self.wb = Workbook()
#             self.wb.remove(self.wb['Sheet'])
#             self.ws = self.wb.create_sheet(title=name_sheet, index=0)
#             # Присваиваю ему новый цвет
#             self.ws.sheet_properties.tabColor = 'DC143C'
#             # Шапка таблицы
#             self.ws.append(['По какое число', 'Имя', 'Цена старая', 'Цена новая', 'Магазин'])
#
#     def save(self, name):
#         filename = '{}.xlsx'.format(name)
#         self.wb.save(filename)
#
#     def cell_value(self, mass):
#         for item in mass:
#             # print(item['date'], item['name'], item['new/old price'], item['shop'])
#             self.ws.append([item['date'], item['name'], item['old_price'], item['new_price'], item['shop']])
#
#     def start(self, mass):
#         self.create('EdadilParse', 'Okay')
#         self.cell_value(mass)
#         self.save("Edadil_Parse")


class Writer2:
    def __init__(self):
        self.wb = None
        self.ws = None
        self.mass = []
        self.sheet_mass = []
        self.filename = ''

    # Проверка на существование файла
    def check_file(self):
        if not os.path.isfile(self.filename):
            print('Create file: {}'.format(self.filename))
            Log().write_log(['info', 'Create file: {}'.format(self.filename)])
            self.wb = Workbook()
            self.wb.remove(self.wb['Sheet'])
            # Сохраняю новую таблицу
        else:
            print('File is exist: {}'.format(self.filename))
            Log().write_log(['info', 'File is exist: {}'.format(self.filename)])
            self.wb = load_workbook(self.filename)

    # Проверка на существование листа
    def check_sheet(self):
        for sheet in self.sheet_mass:
            if sheet in self.wb.sheetnames:
                print('Лист уже есть: {}'.format(sheet))
                Log().write_log(['info', 'Лист уже есть: {}'.format(sheet)])
            else:
                print('Создаю лист: {}'.format(sheet))
                Log().write_log(['info', 'Создаю лист: {}'.format(sheet)])
                self.ws = self.wb.create_sheet(title=sheet)
                self.ws.append(['По какое число', 'Имя', 'Цена старая', 'Цена новая', 'Магазин'])
        self.save(self.filename)
        self.sheet_mass = []

    # Получение массива с данными
    def get_mass(self):
        # Получаем магазины и записываем в массив с названиями листов
        Log().write_log(['info', 'Получаем магазины и записываем в массив с названиями листов'])
        for name_sheet in self.mass:
            if not name_sheet[0] in self.sheet_mass:
                self.sheet_mass.append(name_sheet[0])

    def write_data(self):
        Log().write_log(['info', 'Записываю данные'])
        for item in self.mass:
            for el in item[1:]:
                ws = self.wb.get_sheet_by_name(el[0]['shop'])
                ws.append([el[0]['date'], el[0]['name'], el[0]['old_price'], el[0]['new_price'], el[0]['shop']])
        self.save(self.filename)

    def save(self, filename):
        Log().write_log(['info', 'Сохраняю'])
        self.wb.save(filename)

    def run(self, mass):
        name_file = 'ParsingSite(Eda)'
        path = 'out\\'

        date = datetime.datetime.now().strftime('%d_%m_%Y')

        self.mass = mass
        self.filename = path + '{}_{}.xlsx'.format(name_file, date)
        self.check_file()
        self.get_mass()
        self.check_sheet()
        self.write_data()


